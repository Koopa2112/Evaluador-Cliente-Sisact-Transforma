from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import time
import traceback

# Ruta al controlador de Chrome, ajusta la ruta según donde hayas descargado el controlador
service = Service(executable_path='C:/Chromedriver/chromedriver-win64/chromedriver.exe')
options = webdriver.ChromeOptions()
options.binary_location = "C:/Chromedriver/chrome-win64/chrome.exe"
options.add_argument('--disable-gpu')
options.add_argument('--start-maximized')

# Inicializar el navegador
driver = webdriver.Chrome(service=service, options=options)

#Especificar el tiempo de espera
wait = WebDriverWait(driver, 45)


def accesoPortalDistribuidor():
	# Abrir una página web
	portal = "https://www.distribuidor.telcel.com:4475/Portal-Distribuidores/app/login"
	driver.get(portal)

	#Login Portal Distribuidor
	user = driver.find_element("id", 'j_username')
	user.send_keys('')#<---AQUI VA EL USER
	pwd = driver.find_element("id", 'pwd')
	pwd.send_keys(''). #<--- AQUI VA EL PASSWORD
	region = driver.find_element("id", 'cmbRegiones')
	select = Select(region)
	select.select_by_value("4")
	#time.sleep(1)
	loginPD = driver.find_element("id", 'myBtn')
	loginPD.click()
	time.sleep(3)
	#Dentro de Portal Distribuidor
	texto_del_enlace = "Activaciones"
	enlace_xpath = driver.find_element("xpath", f'//a[text()="{texto_del_enlace}"]')
	enlace_xpath.click()

def accesoSisactTransforma():
	texto_del_span = "SISACT Transforma"
	Sisact = driver.find_element("xpath", f'//a[span[text()="{texto_del_span}"]]')
	driver.execute_script("arguments[0].click();", Sisact)

	time.sleep(3)
	#login de Sisact Transforma
	driver.switch_to.window(driver.window_handles[1])
	while True:
		userST = driver.find_element("id", 'username')
		userST.clear();
		userST.send_keys('')#<---AQUI VA EL USER
		pwdST = driver.find_element("id", 'pass')
		pwdST.send_keys('')#<---AQUI VA EL PASSWRD
		#time.sleep(1)
		loginST = driver.find_element("id", 'btnLogin')
		loginST.click()
		if "https://www.distribuidor.telcel.com:4477/SisactWeb/faces/index.xhtml" in driver.current_url:
			break

def cargaDeBase():
	#Carga de base para scrap
	base = "base.xlsx"
	libro = load_workbook(base)
	hoja = libro.active
	numerosABuscar = [celda.value for celda in hoja["A"] if celda.value is not None]
	print("Numeros a buscar: ",len(numerosABuscar))
	return(numerosABuscar)

def busqueda(numerosABuscar):

	i = 0
	datos =[]
	try:
	#Ciclo a repetir para hacer el scrapping de numeros
		for numero in numerosABuscar:
			try:

				i+=1
				wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")
				tramites = driver.find_element("id", 'idMenuSenal')
				wait.until(EC.element_to_be_clickable((By.ID, 'idMenuSenal')))
				#time.sleep(0.3)

				tramites.click()												#click a el icono de señal

				texto_del_span = "Renovación"
				#esperarAntesDeBuscar = wait.until(EC.presence_of_element_located(By.XPATH, "//*[contains(text(), 'Renovacion')]"))
				time.sleep(2)
				wait.until(EC.presence_of_element_located((By.XPATH, f'//*[@id="formWest:menuWest"]/ul/li[8]/a')))
				Sisact = driver.find_element("xpath", f'//*[@id="formWest:menuWest"]/ul/li[8]/a')
				driver.execute_script("arguments[0].click();", Sisact)



				wait.until(EC.presence_of_element_located((By.ID, "dlgIdentClte")))
				ventana = driver.find_element("id", 'dlgIdentClte')
				n = False
				while n == False:
					try:
						estilo = ventana.value_of_css_property("display")
						if estilo == "block":
							n = True
						time.sleep(1)
					except StaleElementReferenceException:
						ventana = driver.find_element("id", 'dlgIdentClte')
					except  NoSuchElementException:
						ventana = driver.find_element("id", 'dlgIdentClte')
						continue

				wait.until(EC.presence_of_element_located((By.ID, "formIdent:folioid")))
				campoNumero = driver.find_element("id", 'formIdent:folioid')
				print(str(i), "/", str(len(numerosABuscar)), numero)
				campoNumero.send_keys(str(numero))
				#plazo_contrato = driver.find_element("id", 'formIdent:plazoid_input')
				#select = Select(plazo_contrato)
				#select.select_by_value("F")
				buscar = driver.find_element("id", 'formIdent:sig')
				buscar.click()
				espera_avance = wait.until(EC.invisibility_of_element_located((By.ID, "statusDialog")))
				

				#Despues de que se busca el numero en el sistema
				if "No se encontró" in driver.page_source or "sin líneas activas" in driver.page_source or "Error de" in driver.page_source or "Línea Bloqueada" in driver.page_source or "Ocurrio un error" in driver.page_source:
					datos.append((numero, "Error General", "No info",""))
					time.sleep(0.2)
					wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[aria-label='Close']")))
					driver.find_element(By.CSS_SELECTOR, "a[aria-label='Close']").click()

				elif "Cliente encontrado" in driver.page_source:
					espera_boton = wait.until(EC.element_to_be_clickable((By.ID, "formIdent:btnContinuar")))
					#print(driver.find_element("id", 'formIdent:msg').text)
					driver.find_element("id", 'formIdent:btnContinuar').click()
					#print("hola")
					#espera_aparecer_boton = wait.until(EC.presence_of_element_located((By.ID, "j_id_1k_f8:j_id_1k_fa")))
					#try:
					#	print("Esperando")
					#	wait.until(EC.element_to_be_clickable((By.ID, "j_id_1k_j8")))
					#	driver.find_element("id", 'j_id_1k_j8').click()
					#except TimeoutException:
					#	print("en espera")

					#espera_presionar_boton = wait.until(EC.element_to_be_clickable((By.ID, "j_id_1k_f8:j_id_1k_fa")))
					#time.sleep(0.3)
					#wait.until(EC.presence_of_element_located((By.ID, "j_id_1k_g4"))) #Esperar ventana de condiciones de la linea
					wait.until(EC.presence_of_element_located((By.XPATH, f'//*[@id="j_id_1k_ex:j_id_1k_ez"]'))) #Esperar botón de continuar
					monto_restante = driver.find_element("xpath", f'//*[@id="j_id_1k_e8"]').text		#Aqui es la ventana de monto
					plazo_linea = driver.find_element("xpath", f'//*[@id="j_id_1k_ek"]').text		#Aqui es el plazo de linea
					dias_restantes = driver.find_element("xpath", f'//*[@id="j_id_1k_eo"]').text	#para guardar los dias
					print("Restante :$", monto_restante, ", Plazo: " ,plazo_linea, ". Dias restantes: ", dias_restantes)
					#obtengo los meses													restantes
					try:
						wait.until(EC.element_to_be_clickable((By.ID, 'j_id_1k_ex:j_id_1k_ez'))) #click continuar ventana condiciones actuales
						driver.find_element("id", 'j_id_1k_ex:j_id_1k_ez').click()
					except ElementClickInterceptedException:
						time.sleep(0.5)
						wait.until(EC.element_to_be_clickable((By.ID, 'j_id_1k_ex:j_id_1k_ez')))
						driver.find_element("id", 'j_id_1k_ex:j_id_1k_ez').click()
					#Obtengo el nombre del cliente
					nombre_cliente = driver.find_element("id", 'idformnorth:j_id_1h_i').text
					municipio = driver.find_element("id", 'formUsuario:domicilios:0:mun').get_attribute('value')
					try:
						select_element = driver.find_element("id", 'formUsuario:domicilios:0:col_input')
						select = Select(select_element)
						colonia = select.first_selected_option.text
					except NoSuchElementException as e:
						colonia = driver.find_element("id", 'formUsuario:domicilios:0:col_editableInput').get_attribute('value')
					plan = driver.find_element("id", 'formUsuario:j_id_1k_7z').text         #Obtengo el plan

					#tipo_renovacion = driver.find_element("id", 'formUsuario:tipoRen_input')
					#driver.execute_script("arguments[0].scrollIntoView(true);", tipo_renovacion)
					#select = Select(tipo_renovacion)
					#select.select_by_value("11")			#Selecciono renovación por plazo concluido
					#driver.find_element("id", 'formUsuario:sig').click()

					#wait.until(EC.presence_of_element_located((By.ID, "dlgChkForm:j_id_1k_h9_0_4")))
					#adeudo = driver.find_element("id", "dlgChkForm:j_id_1k_h9_0_4").text

					#driver.find_element("xpath", f'//*[@id="dlgCredit1"]/div[1]/a').click()
					
					datos.append((numero, nombre_cliente, municipio, colonia, plazo_linea, dias_restantes, plan, monto_restante))	
						
				else:
					datos.append((numero, "Error General", "No info",""))
					driver.refresh()
			except TimeoutException:
				datos.append((numero, "Revisar manualmente", "Tiempo agotado","TimeoutException"))
				print("Tiempo excedido")
				traceback.print_exc()
				driver.refresh()
				continue
			except ElementClickInterceptedException:
				datos.append((numero, "Revisar de nuevo", "Favor de volver a revisar","ElementClickInterceptedException:"))
				print("No se puede clickear algun elemento")
				traceback.print_exc()
				driver.refresh()
				continue
			#except NoSuchElementException:
			#	print("No se encontro el elemento")
			#	driver.refresh()
			#	continue
			except Exception as e:
				datos.append((numero, "Error", f"Favor de volver a revisar: {e}"  ))
				print(f"Error desconocido en la busqueda: {e}")
				traceback.print_exc()
				driver.delete_all_cookies()
				driver.close()
				driver.switch_to.window(driver.window_handles[0])
				driver.delete_all_cookies()
				accesoPortalDistribuidor()
				accesoSisactTransforma()
				continue
	except Exception as e:
		print(f"Error desconocido general: {e}")
		traceback.print_exc()
	finally:
		return(datos)
	
def guardar(datos):
	#Creacion de libro con la informaciónç
	formato_fecha = "%Y-%m-%d_%H-%M-%S"
	fecha_hora_actual = datetime.now().strftime(formato_fecha)
	scrap = Workbook()
	hoja_scrap = scrap.active
	hoja_scrap.append(["Numero de linea", "Nombre del cliente", "Municipio", "Colonia", "Plazo", "Dias Restantes", "Plan", "Monto"])
	#numero, nombre_cliente, municipio, colonia, plazo_linea, dias_restantes, plan, monto_restante
	for dato in datos:
		hoja_scrap.append(dato)
		scrap.save(f"Base depurada_{fecha_hora_actual}.xlsx")

try:
	accesoPortalDistribuidor()

	accesoSisactTransforma()

	datos = busqueda(cargaDeBase())

except Exception as e:
	print(f"Error desconocido: {e}")
	traceback.print_exc()
finally:
	guardar(datos)


#driver.quit()
